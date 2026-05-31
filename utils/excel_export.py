import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io


def format_minutes(minutes):
    if not minutes:
        return "-"
    h = minutes // 60
    m = minutes % 60
    return f"{h}h {m:02d}m"


def create_excel_report(records, title="Attendance Report"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Styles
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].alignment = center

    # Headers
    headers = ["Employee ID", "Full Name", "Date", "Clock In", "Clock Out", "Worked Hours", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data
    for row_idx, record in enumerate(records, 5):
        data = [
            record.get('employee_id', ''),
            record.get('full_name', ''),
            str(record.get('attendance_date', '')),
            record['clock_in'].strftime('%H:%M') if record.get('clock_in') else '-',
            record['clock_out'].strftime('%H:%M') if record.get('clock_out') else '-',
            format_minutes(record.get('total_minutes')),
            record.get('status', '')
        ]
        fill_color = "F0FFF4" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill("solid", fgColor=fill_color)

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = row_fill
            cell.alignment = center
            cell.border = border

    # Column widths
    widths = [14, 22, 12, 10, 10, 14, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
